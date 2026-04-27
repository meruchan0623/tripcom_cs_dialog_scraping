from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .models import ImportPreview, ImportRoleCount, SessionRecord
from .state import dedupe_sessions


def export_links_xlsx(path: Path, sessions: list[SessionRecord]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    grouped: dict[str, list[SessionRecord]] = defaultdict(list)
    for s in sessions:
        grouped[s.cs_name or "Unknown"].append(s)

    for cs_name, rows in grouped.items():
        sheet_name = (cs_name or "Unknown").replace("/", "_").replace("\\", "_")[:31] or "Unknown"
        ws = wb.create_sheet(sheet_name)
        ws.append(["序号", "会话ID", "客服", "创建时间", "详情页链接"])
        for idx, s in enumerate(rows, start=1):
            ws.append([idx, s.session_id, s.cs_name, s.create_time, s.detail_url])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def import_links_xlsx(path: Path) -> list[SessionRecord]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sessions: list[SessionRecord] = []
    try:
        for ws in wb.worksheets:
            for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if ridx == 1:
                    continue
                if not row:
                    continue
                values = [str(x or "").strip() for x in row]
                if len(values) < 5:
                    continue
                sid = values[1]
                if not sid:
                    continue
                sessions.append(
                    SessionRecord(
                        session_id=sid,
                        cs_name=values[2] or "Unknown",
                        create_time=values[3] or "",
                        detail_url=values[4] or "",
                        imported=True,
                    ).normalized()
                )
    finally:
        wb.close()
    return dedupe_sessions(sessions)


def preview_sessions(sessions: list[SessionRecord]) -> ImportPreview:
    counter = Counter((s.cs_name or "Unknown") for s in sessions)
    roles = [ImportRoleCount(cs_name=name, count=count) for name, count in sorted(counter.items(), key=lambda x: (-x[1], x[0]))]
    return ImportPreview(total_sessions=len(sessions), total_roles=len(roles), roles=roles)

